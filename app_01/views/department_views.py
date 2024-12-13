
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt

from app_01.models import Department


def department_list(request):
    data_list = Department.objects.all()

    from app_01.utils.pagination import Pagination

    my_page = Pagination(request,data_list,page_size=10)


    return render(request, 'department_list.html', {'data_list': my_page.data_list,'page_param':my_page.page_param,'page_str':my_page.page_str()})


def department_add(request):
    if request.method == "GET":
        return render(request, 'department_add.html')

    department_name = request.POST.get('department_name')

    Department.objects.create(department_name=department_name)

    return redirect('/depart/list')


def department_delete(request):
    id = request.GET.get('id')
    Department.objects.get(id=id).delete()

    return redirect('/depart/list')


def department_edit(request, id):
    if request.method == "GET":
        item = Department.objects.filter(id=id).first()

        return render(request, 'department_edit.html', {'item': item})

    department_name = request.POST.get('department_name')
    Department.objects.filter(id=id).update(department_name=department_name)

    return redirect('/depart/list')



@csrf_exempt
def department_multi(request):
    from openpyxl import load_workbook


    file_obj = request.FILES.get('excel_file')

    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows():
        id = row[0].value
        name = row[1].value

        Department.objects.create(pk=id,department_name=name)


    return HttpResponse('uploading')
